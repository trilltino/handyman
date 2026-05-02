"""
Tradesman Cold Call Scraper - FREE Version
Scrapes UK business directories for tradesman contact info.
No API key required!
"""

import re
import time
from datetime import datetime
from pathlib import Path

try:
    import requests
    from bs4 import BeautifulSoup
    from openpyxl import Workbook, load_workbook
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip install openpyxl requests beautifulsoup4 lxml")
    exit(1)

# Configuration
UK_REGIONS = [
    "London",
    "Manchester", 
    "Birmingham",
    "Liverpool",
    "Bristol",
    "Dartford",
    "Coventry",
    "Leeds",
    "Kenilworth",
    "Leamington-Spa",
    "Warwick"
]

TRADE_TYPES = [
    "handyman",
    "plumber",
    "electrician",
    "carpenter",
    "roofer"
]

# Excel file path
OUTPUT_EXCEL = Path(r"C:\Users\isich\XFLeads\XFTradesman Cold Call Document.xlsx")

# Headers to mimic a browser
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-GB,en;q=0.5",
    "Accept-Encoding": "gzip, deflate",
    "Connection": "keep-alive",
}


def scrape_thomsonlocal(trade_type, location, max_pages=2):
    """Scrape Thomson Local for tradesman listings."""
    results = []
    
    # Format for Thomson Local URL
    location_slug = location.lower().replace(" ", "-")
    
    for page in range(1, max_pages + 1):
        try:
            # Thomson Local URL format
            if page == 1:
                url = f"https://www.thomsonlocal.com/search/{trade_type}/{location_slug}"
            else:
                url = f"https://www.thomsonlocal.com/search/{trade_type}/{location_slug}?page={page}"
            
            print(f"  Fetching: {url}")
            
            session = requests.Session()
            response = session.get(url, headers=HEADERS, timeout=20)
            
            if response.status_code != 200:
                print(f"  Status: {response.status_code} - trying alternative...")
                # Try Yelp UK as fallback
                results.extend(scrape_yelp_uk(trade_type, location, 1))
                break
                
            soup = BeautifulSoup(response.text, "lxml")
            
            # Find listings by common patterns
            found_any = False
            
            # Pattern 1: Look for tel: links
            tel_links = soup.find_all("a", href=re.compile(r"tel:"))
            for link in tel_links:
                phone = link.get("href", "").replace("tel:", "").strip()
                if phone and len(re.sub(r'[^\d]', '', phone)) >= 10:
                    # Try to find business name nearby
                    parent = link.find_parent(["div", "article", "li", "section"])
                    if parent:
                        # Look for heading or strong text
                        name_elem = parent.find(["h2", "h3", "h4", "strong", "b"])
                        if name_elem:
                            name = name_elem.get_text(strip=True)
                            if name and len(name) > 2:
                                results.append({
                                    "name": name[:100],  # Limit name length
                                    "phone": phone,
                                    "trade_type": trade_type,
                                    "location": location,
                                    "scraped_date": datetime.now().strftime("%Y-%m-%d %H:%M")
                                })
                                print(f"    ✓ {name[:40]} - {phone}")
                                found_any = True
            
            if not found_any:
                print(f"  No results on page {page}")
                break
                
            time.sleep(1.5)
            
        except Exception as e:
            print(f"  Error: {e}")
            break
    
    return results


def scrape_yelp_uk(trade_type, location, max_pages=1):
    """Scrape Yelp UK as a fallback."""
    results = []
    
    try:
        # Yelp UK search URL
        url = f"https://www.yelp.co.uk/search?find_desc={trade_type}&find_loc={location}%2C+UK"
        
        print(f"  Trying Yelp UK...")
        response = requests.get(url, headers=HEADERS, timeout=15)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, "lxml")
            
            # Look for phone numbers in the page
            phone_pattern = re.compile(r'0\d{2,4}\s?\d{3,4}\s?\d{3,4}')
            text = soup.get_text()
            phones = phone_pattern.findall(text)
            
            for phone in phones[:5]:  # Limit to 5 per page
                results.append({
                    "name": f"{trade_type.title()} in {location}",
                    "phone": phone,
                    "trade_type": trade_type,
                    "location": location,
                    "scraped_date": datetime.now().strftime("%Y-%m-%d %H:%M")
                })
                
    except Exception as e:
        print(f"  Yelp error: {e}")
    
    return results


def scrape_118118(trade_type, location, max_pages=1):
    """Scrape 118118.com for business listings."""
    results = []
    
    try:
        url = f"https://www.118118.com/{trade_type}/{location}"
        print(f"  Trying 118118.com...")
        
        response = requests.get(url, headers=HEADERS, timeout=15)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, "lxml")
            
            # Look for business cards
            listings = soup.find_all("div", class_=re.compile(r"result|listing|business", re.I))
            
            for listing in listings[:10]:
                name_elem = listing.find(["h2", "h3", "a"])
                phone_elem = listing.find("a", href=re.compile(r"tel:"))
                
                if phone_elem:
                    phone = phone_elem.get("href", "").replace("tel:", "").strip()
                    name = name_elem.get_text(strip=True) if name_elem else f"{trade_type.title()} Service"
                    
                    if phone and len(re.sub(r'[^\d]', '', phone)) >= 10:
                        results.append({
                            "name": name[:100],
                            "phone": phone,
                            "trade_type": trade_type,
                            "location": location,
                            "scraped_date": datetime.now().strftime("%Y-%m-%d %H:%M")
                        })
                        print(f"    ✓ {name[:40]} - {phone}")
                        
    except Exception as e:
        print(f"  118118 error: {e}")
    
    return results


def scrape_all_sources(trade_type, location):
    """Try multiple sources to gather leads."""
    all_results = []
    
    # Try Thomson Local first
    results = scrape_thomsonlocal(trade_type, location, max_pages=2)
    all_results.extend(results)
    
    # Try 118118 as backup
    if len(results) < 3:
        results = scrape_118118(trade_type, location)
        all_results.extend(results)
    
    return all_results


def setup_workbook(output_path):
    """Create or load Excel workbook."""
    if output_path.exists():
        print(f"Loading existing workbook: {output_path}")
        wb = load_workbook(output_path)
        ws = wb.active
    else:
        print("Creating new workbook...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        headers = ["Business Name", "Phone Number", "Trade Type", "Location", "Scraped Date", "Called", "Notes"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
    
    return wb, ws


def get_existing_phones(ws):
    """Get set of phone numbers already in sheet."""
    existing = set()
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[1]:
            phone = re.sub(r'[^\d]', '', str(row[1]))
            existing.add(phone)
    return existing


def add_to_excel(ws, businesses, existing_phones):
    """Add new businesses to Excel."""
    next_row = ws.max_row + 1
    if next_row == 2 and ws.cell(row=1, column=1).value is None:
        headers = ["Business Name", "Phone Number", "Trade Type", "Location", "Scraped Date", "Called", "Notes"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
    
    added = 0
    for biz in businesses:
        phone_normalized = re.sub(r'[^\d]', '', str(biz["phone"]))
        
        if phone_normalized not in existing_phones and len(phone_normalized) >= 10:
            ws.cell(row=next_row, column=1, value=biz["name"])
            ws.cell(row=next_row, column=2, value=biz["phone"])
            ws.cell(row=next_row, column=3, value=biz["trade_type"])
            ws.cell(row=next_row, column=4, value=biz["location"])
            ws.cell(row=next_row, column=5, value=biz["scraped_date"])
            ws.cell(row=next_row, column=6, value="No")
            ws.cell(row=next_row, column=7, value="")
            
            existing_phones.add(phone_normalized)
            next_row += 1
            added += 1
    
    return added


def main():
    print("=" * 60)
    print("TRADESMAN COLD CALL SCRAPER - FREE VERSION")
    print("=" * 60)
    print("\nNo API key required! Scraping UK directories...")
    
    print(f"\nRegions: {len(UK_REGIONS)}")
    for region in UK_REGIONS:
        print(f"  • {region}")
    
    print(f"\nTrades: {', '.join(TRADE_TYPES)}")
    
    # Setup Excel
    wb, ws = setup_workbook(OUTPUT_EXCEL)
    existing_phones = get_existing_phones(ws)
    print(f"\n✓ Workbook loaded with {len(existing_phones)} existing entries")
    
    # Scrape
    all_businesses = []
    total = len(UK_REGIONS) * len(TRADE_TYPES)
    current = 0
    
    for region in UK_REGIONS:
        for trade in TRADE_TYPES:
            current += 1
            print(f"\n[{current}/{total}] {trade} in {region}")
            
            businesses = scrape_all_sources(trade, region)
            all_businesses.extend(businesses)
            print(f"  Found: {len(businesses)} contacts")
            
            time.sleep(0.5)
    
    # Save to Excel
    print(f"\n\nSaving {len(all_businesses)} entries to Excel...")
    added = add_to_excel(ws, all_businesses, existing_phones)
    wb.save(OUTPUT_EXCEL)
    
    print(f"\n{'=' * 60}")
    print("✓ COMPLETE!")
    print(f"  Total found: {len(all_businesses)}")
    print(f"  New added: {added}")
    print(f"  Saved to: {OUTPUT_EXCEL}")
    print("=" * 60)


if __name__ == "__main__":
    main()
